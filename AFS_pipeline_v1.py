#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automated Filter Suite - AFS_pipeline_v1
"""
import os, re, json, logging, tempfile, argparse
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple
from pathlib import Path
import pandas as pd
import math

logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
YN = lambda b: 'Yes' if bool(b) else 'No'

@dataclass
class Match:
    start: int; end: int; text: str; pattern: str

def ensure_dir(p: str):
    if p and not os.path.exists(p): os.makedirs(p, exist_ok=True)

def load_json(p: str) -> Dict[str, Any]:
    with open(p,'r',encoding='utf-8') as f: return json.load(f)

def validate_patterns_schema(p: Dict[str, Any]) -> List[str]:
    errs: List[str] = []
    if not isinstance(p, dict):
        return ["Top-level patterns object must be a JSON object."]
    if 'regex_sets' not in p or not isinstance(p['regex_sets'], dict):
        errs.append("Missing or invalid 'regex_sets' (expected object).")
    if 'gating' in p and not isinstance(p['gating'], dict):
        errs.append("'gating' must be an object.")
    if 'brand_disambiguation' in p and not isinstance(p['brand_disambiguation'], dict):
        errs.append("'brand_disambiguation' must be an object.")
    if 'blocklists' in p and not isinstance(p['blocklists'], dict):
        errs.append("'blocklists' must be an object.")
    return errs

def regex_hygiene_scan(patterns: Dict[str, Any]) -> List[Dict[str,str]]:
    rows: List[Dict[str,str]] = []
    rs = patterns.get('regex_sets', {}) or {}
    for set_name, pats in rs.items():
        for pat in pats or []:
            issues = []
            if pat.count('(') != pat.count(')'):
                issues.append('Unbalanced parentheses')
            if pat.endswith('\\'):
                issues.append('Trailing backslash')
            if re.search(r'(?<!\\)\.(?!\*)', pat):
                issues.append("Unescaped dot '.'")
            if not pat.strip():
                issues.append('Empty pattern')
            if issues:
                rows.append({'Set': set_name, 'Pattern': pat, 'Issue': '; '.join(issues)})
    return rows

def compile_regex_set(patterns: List[str], set_name: str, issues: List[Dict[str,str]]) -> List[re.Pattern]:
    out: List[re.Pattern] = []
    for ptn in patterns or []:
        try:
            out.append(re.compile(ptn, flags=re.IGNORECASE))
        except re.error as e:
            issues.append({'Set': set_name, 'Pattern': ptn, 'Issue': f'Compile error: {e}'})
    return out

def apply_brand_disambiguation(text: str, hits: List[Match], disamb: Dict[str, Any]) -> List[Match]:
    if not disamb: return hits
    size = int((disamb.get('window') or {}).get('size', 40))
    rules = disamb.get('rules') or []
    kept: List[Match] = []
    for h in hits:
        token = h.text.strip().lower()
        rule = None
        for r in rules:
            tok = (r.get('token') or '').strip().lower()
            if tok and tok == token:
                rule = r; break
        if rule is None:
            kept.append(h); continue
        left = max(0, h.start - size)
        right = min(len(text), h.end + size)
        ctx = text[left:right].lower()
        pos = [a.lower() for a in (rule.get('positive_anchors') or [])]
        neg = [a.lower() for a in (rule.get('negative_contexts') or [])]
        pos_ok = True if not pos else any(a in ctx for a in pos)
        neg_ok = not any(a in ctx for a in neg)
        if pos_ok and neg_ok: kept.append(h)
    return kept

def positions(text: str, pats: List[re.Pattern]) -> List[Tuple[int, int]]:
    pos: List[Tuple[int, int]] = []
    for rx in pats:
        for m in rx.finditer(text): pos.append((m.start(), m.end()))
    return sorted(pos)

def link_hit_positions(a_pos: List[Tuple[int, int]], b_pos: List[Tuple[int, int]], window: int) -> bool:
    i=j=0
    while i<len(a_pos) and j<len(b_pos):
        a=a_pos[i]; b=b_pos[j]
        dist=min(abs(a[0]-b[1]),abs(b[0]-a[1]),abs(a[0]-b[0]),abs(a[1]-b[1]))
        if dist<=window: return True
        if a[1]<b[1]: i+=1
        else: j+=1
    return False

def adapt_patterns_for_lit(patterns: Dict[str, Any]) -> Dict[str, Any]:
    if 'regex_sets' not in patterns:
        raise ValueError("Expected v2 pattern schema with 'regex_sets'.")
    rs = patterns['regex_sets'] or {}
    g_in = (patterns.get('gating') or {})
    def gval(key, default):
        return g_in.get(key, patterns.get(key, default))
    g = {
        'Org_flag_requires_AI_proximity': bool(gval('Org_flag_requires_AI_proximity', True)),
        'Org_AI_proximity_window': int(gval('Org_AI_proximity_window', 50)),
        'Regulation_flag_requires_AI_proximity': bool(gval('Regulation_flag_requires_AI_proximity', True)),
        'Reg_AI_proximity_window': int(gval('Reg_AI_proximity_window', 75)),
        'Reg_Org_proximity_window': int(gval('Reg_Org_proximity_window', 75)),
    }
    out: Dict[str, Any] = {
        'AI': rs.get('AI', []),
        'GenAI_strict': rs.get('GenAI_Brands_Strict', []) or [],
        'GenAI_ambig': rs.get('GenAI_Brands_Ambiguous', []) or [],
        'Org': rs.get('Org', []),
        'Regulation': rs.get('Regulation', []),
        'Actor': rs.get('Actor', []),
        'Logic': rs.get('Logic', []),
        'Logic_Governance': rs.get('Logic_Governance', []),
        'Logic_Technology': rs.get('Logic_Technology', []),
        'Logic_Market': rs.get('Logic_Market', []),
        'TechOnly': (((patterns.get('blocklists', {}) or {}).get('TechOnly', {}) or {}).get('phrases', [])),
        'brand_disambiguation': patterns.get('brand_disambiguation', {}) or {},
        'gating': g,
    }
    return out

def litfilter_run(input_path: str, patterns_path: str, output_xlsx: str = None,
                  id_col: str = 'ID', title_col: str = 'Title', abstract_col: str = 'Abstract',
                  sheet_name_hint: str = None, timestamp: bool = False) -> str:
    sheet_name = 'Master'
    tmp_excel = None
    if input_path.lower().endswith('.csv'):
        df = pd.read_csv(input_path)
        tmp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
        with pd.ExcelWriter(tmp_excel, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name=sheet_name)
        input_xlsx = tmp_excel
        input_sheet = sheet_name
    else:
        input_xlsx = input_path
        xls = pd.ExcelFile(input_xlsx, engine='openpyxl')
        input_sheet = sheet_name_hint or (xls.sheet_names[0] if xls.sheet_names else sheet_name)
        df = xls.parse(input_sheet, engine='openpyxl')

    patterns = load_json(patterns_path)
    schema_errs = validate_patterns_schema(patterns)
    if schema_errs:
        raise ValueError("Pattern schema errors: " + "; ".join(schema_errs))

    hygiene_rows = regex_hygiene_scan(patterns)
    ap = adapt_patterns_for_lit(patterns)
    issues: List[Dict[str,str]] = []

    AI = compile_regex_set(ap['AI'], 'AI', issues)
    GenS = compile_regex_set(ap['GenAI_strict'], 'GenAI_strict', issues)
    GenA = compile_regex_set(ap['GenAI_ambig'], 'GenAI_ambig', issues)
    Org = compile_regex_set(ap['Org'], 'Org', issues)
    Reg = compile_regex_set(ap['Regulation'], 'Regulation', issues)
    Actor = compile_regex_set(ap['Actor'], 'Actor', issues)
    Logic = compile_regex_set(ap['Logic'], 'Logic', issues)
    Logic_Gov = compile_regex_set(ap.get('Logic_Governance', []), 'Logic_Governance', issues)
    Logic_Tech = compile_regex_set(ap.get('Logic_Technology', []), 'Logic_Technology', issues)
    Logic_Market = compile_regex_set(ap.get('Logic_Market', []), 'Logic_Market', issues)

    tech_re = None
    try:
        if ap['TechOnly']:
            tech_re = re.compile('|'.join(rf"\\b{re.escape(p)}\\b" for p in ap['TechOnly']), flags=re.IGNORECASE)
    except re.error as e:
        issues.append({'Set':'TechOnly','Pattern':'<joined terms>','Issue':f'Compile error: {e}'})
        tech_re = None

    df_out = df.copy()
    cols_lower = {c.lower(): c for c in df_out.columns}
    title_col = cols_lower.get(title_col.lower(), title_col)
    abstract_col = cols_lower.get(abstract_col.lower(), abstract_col)
    if title_col not in df_out.columns: df_out[title_col] = ''
    if abstract_col not in df_out.columns: df_out[abstract_col] = ''
    if id_col in df_out.columns:
        df_out[id_col] = df_out[id_col].astype(str)
    else:
        logging.warning(f"ID column '{id_col}' not found; continuing without it.")

    titles = df_out[title_col].fillna('').astype(str)
    abstracts = df_out[abstract_col].fillna('').astype(str)

    AI_flag=[]; GenAI_Brand_flag=[]; Org_flag=[]; Regulation_flag=[]
    AI_Reg_proximity=[]; AI_Org_proximity=[]; Reg_Org_proximity=[]
    Gov_Tech=[]; Tech_Market=[]; Gov_Market=[]
    Actor_Logic_gating=[]
    Logic_Governance_flag=[]; Logic_Technology_flag=[]; Logic_Market_flag=[]
    Decision=[]; Explain=[]; Matched_Keywords=[]

    for i in range(len(df_out)):
        t = (titles.iloc[i] or '') + '\n' + (abstracts.iloc[i] or '')
        ai_pos = positions(t, AI)
        genS_hits = [Match(m.start(), m.end(), t[m.start():m.end()], rx.pattern) for rx in GenS for m in rx.finditer(t)]
        genA_raw = [Match(m.start(), m.end(), t[m.start():m.end()], rx.pattern) for rx in GenA for m in rx.finditer(t)]
        genA_hits = apply_brand_disambiguation(t, genA_raw, ap['brand_disambiguation'])
        ai_pos += [(m.start, m.end) for m in (genS_hits + genA_hits)]

        org_pos = positions(t, Org)
        reg_pos = positions(t, Reg)
        actor_any = any(rx.search(t) for rx in Actor)
        logic_any = any(rx.search(t) for rx in Logic)
        logic_gov_any = any(rx.search(t) for rx in Logic_Gov) if Logic_Gov else False
        logic_tech_any = any(rx.search(t) for rx in Logic_Tech) if Logic_Tech else False
        logic_market_any = any(rx.search(t) for rx in Logic_Market) if Logic_Market else False

        ai_hit = bool(ai_pos)
        gen_hit = bool(genS_hits or genA_hits)

        reg_window = ap['gating'].get('Reg_AI_proximity_window', 50)
        org_window = ap['gating'].get('Org_AI_proximity_window', 50)
        reg_org_window = ap['gating'].get('Reg_Org_proximity_window', 75)

        ai_reg_prox = link_hit_positions(sorted(ai_pos), sorted(reg_pos), reg_window) if ai_pos and reg_pos else False
        ai_org_prox = link_hit_positions(sorted(ai_pos), sorted(org_pos), org_window) if ai_pos and org_pos else False
        reg_org_prox = link_hit_positions(sorted(reg_pos), sorted(org_pos), reg_org_window) if reg_pos and org_pos else False

        if ap['gating'].get('Org_flag_requires_AI_proximity', True):
            org_hit = bool(org_pos) and ai_org_prox
        else:
            org_hit = bool(org_pos)

        if ap['gating'].get('Regulation_flag_requires_AI_proximity', True):
            reg_hit = bool(reg_pos) and ai_reg_prox
        else:
            reg_hit = bool(reg_pos)

        if tech_re is not None:
            tech_hits = [m.group(0).lower() for m in tech_re.finditer(t)]
        else:
            tl = t.lower(); tech_hits = [ph for ph in ap['TechOnly'] if ph.lower() in tl]
        tech_only = bool(tech_hits) and not (ai_org_prox or ai_reg_prox or (actor_any and logic_any))

        decision = 'Exclude'
        if not any([ai_hit, gen_hit, org_hit, reg_hit, actor_any, logic_any]):
            decision = 'Exclude (X0)'
        elif tech_only and not reg_hit:
            decision = 'Exclude'
        elif ai_hit and (reg_hit and ai_reg_prox):
            decision = 'Keep'
        elif ai_hit and (actor_any or logic_any) and (ai_org_prox or ai_reg_prox):
            decision = 'Keep'
        elif ai_hit and (actor_any or logic_any):
            decision = 'Maybe' # stricter rule: Actor/Logic without proximity is not enough

        reasons = []
        if ai_hit: reasons.append('AI/GenAI present')
        if gen_hit: reasons.append('GenAI brand present')
        if reg_hit: reasons.append('Regulation present' + (' (AI-prox)' if ai_reg_prox else ''))
        if org_hit: reasons.append('Org present' + (' (AI-prox)' if ai_org_prox else ''))
        if actor_any: reasons.append('Actor present')
        if logic_any: reasons.append('Logic present')
        if logic_gov_any: reasons.append('Logic_Governance present')
        if logic_tech_any: reasons.append('Logic_Technology present')
        if logic_market_any: reasons.append('Logic_Market present')
        if reg_org_prox: reasons.append('Reg–Org proximity')
        if tech_only: reasons.append('Tech-only heuristics')

        explain = ' | '.join(sorted(reasons)) if reasons else 'No evidence.'

        def collect_terms(rx_list: List[re.Pattern]) -> List[str]:
            s = set()
            for rx in rx_list:
                for m in rx.finditer(t): s.add(m.group(0))
            return sorted(s)

        ai_terms = collect_terms(AI)
        gen_terms = sorted({h.text for h in (genS_hits + genA_hits)})
        org_terms = collect_terms(Org)
        reg_terms = collect_terms(Reg)
        actor_terms = collect_terms(Actor)
        logic_terms = collect_terms(Logic)
        logic_gov_terms = collect_terms(Logic_Gov) if Logic_Gov else []
        logic_tech_terms = collect_terms(Logic_Tech) if Logic_Tech else []
        logic_market_terms = collect_terms(Logic_Market) if Logic_Market else []

        buckets = []
        if ai_terms: buckets.append('AI: ' + '; '.join(sorted(set(ai_terms))))
        if gen_terms: buckets.append('GenAI_Brand: ' + '; '.join(sorted(set(gen_terms))))
        if org_terms: buckets.append('Org: ' + '; '.join(sorted(set(org_terms))))
        if reg_terms: buckets.append('Regulation: ' + '; '.join(sorted(set(reg_terms))))
        if actor_terms: buckets.append('Actor: ' + '; '.join(sorted(set(actor_terms))))
        if logic_terms: buckets.append('Logic: ' + '; '.join(sorted(set(logic_terms))))
        if logic_gov_terms: buckets.append('Logic_Governance: ' + '; '.join(sorted(set(logic_gov_terms))))
        if logic_tech_terms: buckets.append('Logic_Technology: ' + '; '.join(sorted(set(logic_tech_terms))))
        if logic_market_terms: buckets.append('Logic_Market: ' + '; '.join(sorted(set(logic_market_terms))))
        if tech_hits: buckets.append('TechOnly: ' + '; '.join(sorted(set(tech_hits))))
        matched_str = ' \n'.join(buckets)[:32000]

        AI_flag.append(YN(ai_hit))
        GenAI_Brand_flag.append(YN(gen_hit))
        Org_flag.append(YN(org_hit))
        Regulation_flag.append(YN(reg_hit))
        AI_Reg_proximity.append(YN(ai_reg_prox))
        AI_Org_proximity.append(YN(ai_org_prox))
        Reg_Org_proximity.append(YN(reg_org_prox))
        Gov_Tech.append(YN(ai_reg_prox))
        Tech_Market.append(YN(ai_org_prox))
        Gov_Market.append(YN(reg_org_prox))
        Actor_Logic_gating.append(YN(actor_any and logic_any))
        Logic_Governance_flag.append(YN(logic_gov_any))
        Logic_Technology_flag.append(YN(logic_tech_any))
        Logic_Market_flag.append(YN(logic_market_any))
        Decision.append(decision)
        Explain.append(explain)
        Matched_Keywords.append(matched_str)

    df_out['AI_flag'] = AI_flag
    df_out['GenAI_Brand_flag'] = GenAI_Brand_flag
    df_out['Org_flag'] = Org_flag
    df_out['Regulation_flag'] = Regulation_flag
    df_out['AI_Reg_proximity'] = AI_Reg_proximity
    df_out['AI_Org_proximity'] = AI_Org_proximity
    df_out['Reg_Org_proximity'] = Reg_Org_proximity
    df_out['Gov_Tech'] = Gov_Tech
    df_out['Tech_Market'] = Tech_Market
    df_out['Gov_Market'] = Gov_Market
    df_out['Actor_Logic_gating'] = Actor_Logic_gating
    df_out['Logic_Governance'] = Logic_Governance_flag
    df_out['Logic_Technology'] = Logic_Technology_flag
    df_out['Logic_Market'] = Logic_Market_flag
    df_out['Decision'] = Decision
    df_out['Explain'] = Explain
    df_out['Matched_Keywords'] = Matched_Keywords

    total = len(df_out)
    def pct(n): return round((n/total)*100.0, 2) if total else 0.0
    def count_yes(col): return int((df_out[col] == 'Yes').sum())

    decisions = ['Keep','Maybe','Exclude','Exclude (X0)']
    rows = []

    # COLUMN-normalized composition (vertical)
    keys = ['Gov_Tech','Tech_Market','Gov_Market','Logic_Governance','Logic_Technology','Logic_Market']
    total_yes = {k: int((df_out[k] == 'Yes').sum()) for k in keys}
    counts = {lab: {k: int(((df_out['Decision'] == lab) & (df_out[k] == 'Yes')).sum()) for k in keys} for lab in decisions}

    def column_round(vals):
        floored = [math.floor(x * 100) / 100.0 for x in vals]
        remainder = [x - f for x, f in zip(vals, floored)]
        needed = int(round((100.0 - sum(floored)) * 100))
        order = sorted(range(len(remainder)), key=lambda i: remainder[i], reverse=True)
        alloc = [0] * len(remainder)
        for idx in order[:needed]:
            alloc[idx] += 1
        return [round(f + (a/100.0), 2) for f, a in zip(floored, alloc)]

    col_perc = {lab: {} for lab in decisions}
    for k in keys:
        tot = total_yes[k]
        if tot == 0:
            for lab in decisions:
                col_perc[lab][k] = 0.0
            continue
        raw = [counts[lab][k] * 100.0 / tot for lab in decisions]
        rounded = column_round(raw)
        for lab, val in zip(decisions, rounded):
            col_perc[lab][k] = val

    for lab in decisions:
        mask = (df_out['Decision'] == lab)
        n = int(mask.sum())
        row = {
            'Metric': lab,
            'Count': n,
            'Percent': pct(n),
            'Gov_Tech_%': col_perc[lab]['Gov_Tech'],
            'Tech_Market_%': col_perc[lab]['Tech_Market'],
            'Gov_Market_%': col_perc[lab]['Gov_Market'],
            'Logic_Governance_%': col_perc[lab]['Logic_Governance'],
            'Logic_Technology_%': col_perc[lab]['Logic_Technology'],
            'Logic_Market_%': col_perc[lab]['Logic_Market'],
        }
        rows.append(row)

    def blank_row(metric_label, n):
        return {
            'Metric': metric_label,
            'Count': n,
            'Percent': pct(n),
            'Gov_Tech_%': '',
            'Tech_Market_%': '',
            'Gov_Market_%': '',
            'Logic_Governance_%': '',
            'Logic_Technology_%': '',
            'Logic_Market_%': '',
        }

    for lab in ['AI_flag','GenAI_Brand_flag','Org_flag','Regulation_flag',
                'AI_Reg_proximity','AI_Org_proximity','Reg_Org_proximity',
                'Gov_Tech','Tech_Market','Gov_Market',
                'Logic_Governance','Logic_Technology','Logic_Market',
                'Actor_Logic_gating']:
        n = count_yes(lab)
        rows.append(blank_row(lab, n))

    summary_df = pd.DataFrame(rows, columns=['Metric','Count','Percent',
                                             'Gov_Tech_%','Tech_Market_%','Gov_Market_%',
                                             'Logic_Governance_%','Logic_Technology_%','Logic_Market_%'])
    # Hygiene sheet
    hygiene_df = pd.DataFrame(hygiene_rows + issues, columns=['Set','Pattern','Issue']) \
    if (hygiene_rows or issues) else pd.DataFrame(columns=['Set','Pattern','Issue'])

    # Write output and add footnotes to Summary
    out_path = './outputs/_AFS_LitFilter_v1.xlsx' if output_xlsx is None else output_xlsx
    ensure_dir(os.path.dirname(out_path))
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name=input_sheet)
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        hygiene_df.to_excel(writer, index=False, sheet_name='Regex_Validator')

        # Footnotes
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
        ws = writer.sheets['Summary']
        start_row = summary_df.shape[0] + 3
        title = "Notes (column-normalized view)"
        bullets = [
            "• The six *_% columns are vertical compositions: within each indicator, percentages across Keep/Maybe/Exclude/Exclude (X0) sum to 100%.",
            "• Example: Gov_Tech_% shows how often Gov_Tech=Yes falls into each decision bucket.",
            "• If an indicator has no 'Yes' in the corpus, its entire column is 0.00%.",
            "• Indicators can co-occur; these columns are not meant to sum horizontally per row.",
            "• 'Percent' column (beside Count) is the share of the entire corpus in that decision, independent of the six *_% columns."
        ]
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
        for i, line in enumerate(bullets, start=1):
            cell = ws.cell(row=start_row + i, column=1, value=line)
            cell.alignment = Alignment(wrap_text=True)
        # Header wrap and widths
        for col_idx in range(1, 10):
            ws.cell(row=1, column=col_idx).alignment = Alignment(wrap_text=True)
        widths = {1: 22, 2: 10, 3: 10, 4: 14, 5: 16, 6: 14, 7: 18, 8: 18, 9: 16}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'A2'

    if tmp_excel is not None:
        try: os.remove(tmp_excel)
        except Exception: pass
    return out_path


def main():
    p = argparse.ArgumentParser(description='Automated Filter Suite - AFS_pipeline (column-normalized Summary + footnotes)')
    p.add_argument('--input', required=True)
    p.add_argument('--patterns', required=True)
    p.add_argument('--output', default=None)
    p.add_argument('--sheet', default=None)
    p.add_argument('--title-col', default='Title')
    p.add_argument('--abstract-col', default='Abstract')
    p.add_argument('--id-col', default='ID')
    p.add_argument('--timestamp', action='store_true')
    args = p.parse_args()
    out = litfilter_run(
        input_path=args.input,
        patterns_path=args.patterns,
        output_xlsx=args.output,
        id_col=args.id_col,
        title_col=args.title_col,
        abstract_col=args.abstract_col,
        sheet_name_hint=args.sheet,
        timestamp=args.timestamp,
    )
    print(out)

if __name__ == '__main__':
    main()
